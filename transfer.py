import base64
import zipfile
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference, Series
import pandas as pd
from io import BytesIO
import streamlit as st
from shutil import copyfile
import os
from openpyxl import load_workbook, Workbook
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import requests
import time
from datetime import date
import requests
import random
import json
from hashlib import md5
import numpy as np
from PIL import Image
import xlsxwriter
import email.mime.multipart
from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import requests
import time
from datetime import date
today = date.today()
import pandas as pd
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
align = Alignment(horizontal='left', vertical='center')
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)
date_now = time.strftime("%d/%m/%Y", time.localtime())

def translate_eng_cn(query):
    # Set your own appid/appkey.
    appid = '20220629001259722'
    appkey = 'vkooiwx4xLqOl9C8NjvW'
    # For list of language codes, please refer to `https://api.fanyi.baidu.com/doc/21`
    from_lang = 'en'
    to_lang = 'zh'
    endpoint = 'http://api.fanyi.baidu.com'
    path = '/api/trans/vip/translate'
    url = endpoint + path
    # Generate salt and sign
    def make_md5(s, encoding='utf-8'):
        return md5(s.encode(encoding)).hexdigest()
    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()
    description_en_chinois = result['trans_result'][0]['dst']
    return description_en_chinois

def get_data(handler, date_noa, date_pick_up, lta, pcs, kg):  # 海关邮件正文
    global dfges
    dfges = pd.DataFrame([["MTD De Depart", handler],
                          ["MTD D'Arrivee", date_noa],
                          ["Representant douane", "Alando"],
                          ["La date du jour (du transfert)", date_pick_up],
                          ["La date de MDT (Handler de départ)", date_pick_up],
                          ["Lieu de Presentation", "Alando"],
                          ["Le numéro de colis (tracking)", lta],
                          ["L’identification de la marchandise", "/"],
                          ["Le colisage (nombre de colis)", str(pcs) + " PCS"],
                          ["Le kG", str(kg) + " KG"]],
                         columns=['DESCRIPTION', 'INFORMATION'])
    return dfges
def extrait_hscode(hscode, today):
    dic = []
    url = "https://eservices.minfin.fgov.be/extTariffBrowser/Measure?cnCode=%s&country=29422&trade=0&cssfile=tarbro" \
          "&date=%s&lang=EN&page=1" % (
              hscode, today)
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/76.0.3809.132 Safari/537.36'}
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'html.parser')
    hscode = soup.find('span', class_="smaller-title").text.replace(" ","")
    description_hscode = soup.find('ul', class_="nostyle").getText().replace('\n', '').replace(
        '                                   ', ' /')
    footnote = soup.find('table', class_="table-nopadding").getText().replace('\n', '').strip().replace("Footnotes:",
                                                                                                        "")
    supplementary_unit = soup.find('table', class_="table-nopadding bottom-aligned").getText().replace('\n',
                                                                                                       '').strip().replace(
        "Supplementary unit:", "")
    tables = soup.find_all('div', class_="meas-header")
    for table in tables:
        table_infos = table.getText().split("\n")
        type_table = table_infos[0]  # 表格类型
        table_infos_sorts = table_infos[19:]
        nb_ligne = len(table_infos_sorts) / 20
        for x in range(int(nb_ligne)):
            Geographical_area = table_infos_sorts[20 * x + 0] + "  " + table_infos_sorts[20 * x + 1]
            Measure_type = table_infos_sorts[20 * x + 2] + "  " + table_infos_sorts[20 * x + 3]
            Tariff = table_infos_sorts[20 * x + 4] + "  " + table_infos_sorts[20 * x + 5]
            dic_0 = {"type_table":type_table,
                     "Measure_type":Measure_type,
                     "Tariff":Tariff,
            "Geographical_area": Geographical_area,}
            dic.append(dic_0)
    pd_hscode_no_info = pd.DataFrame(list(dic))
    if "CN - China  " in pd_hscode_no_info["Geographical_area"].tolist():
        anti_dumping = "anti-dumping"
    else:
        anti_dumping = ""
    duty = pd_hscode_no_info["Tariff"].loc[(pd_hscode_no_info["type_table"]=="Tariff measures") & (pd_hscode_no_info["Measure_type"]=="Third country duty          ") ].tolist()[0]
    return description_hscode,anti_dumping,duty



def intro():
    import streamlit as st
    st.write("# 欢迎使用SMDG线上服务 👋")
    st.sidebar.success("请选择极致服务")
    st.markdown(
        """
        SMDG Logistics SRL 是一家位于比利时列日，专注于中欧货运服务的跨境物流公司 

        **👈 从左边下拉框，请选择您需要的服务** 去体验SMDG极致的智能化服务
        ### 想要了解更多?
        - 欧盟海关码查询网站  [Tarbel](https://eservices.minfin.fgov.be/extTariffBrowser/browseNomen.xhtml?suffix=80&country=&lang=EN&page=1&date=20220727)
        - 欧盟进出口数据查询网站 [Import - Export Statistics](https://trade.ec.europa.eu/access-to-markets/en/statistics)
        - VAT有效性查询 [VIES VAT Validation](https://ec.europa.eu/taxation_customs/vies/)
        - EORI有效性查询 [Eori Validation](https://ec.europa.eu/taxation_customs/dds2/eos/eori_validation.jsp)
        ### SMDG 尾程派送优质服务
        - 零担    派送 : 涵盖 包括德国、法国、西班牙在内的主要西欧国家的零担派送业务
        - DHL   Paket : 价格低廉，时效快 (一周5天包车派送)
        - UPS     德国 ：覆盖全欧的UPS 快递派送业务
        - 比利时DPD直送 ：CDG7 / DTM2 亚马逊快递直送业务
        - AMM -  WRO5 : 集装箱直送业务
        ### 联系我们
        - 邮箱 : info@smdg.eu
        - 地址 : Rue Louis Bleriot 5A 4460 Bierset Belgiumt
        """)

def custom_invoice():
    import streamlit as st
    import time
    import numpy as np
    st.markdown(f"# 智能生成{list(page_names_to_funcs.keys())[1]}")
    st.write(
        """ 在这里您可以通过上传清关资料线上生成相关的CI和PL """)
    st.subheader("第一步骤 ：上传发票抬头信息")
    invoice_tete = st.file_uploader("", type=(["xlsx", "xls"]))
    if invoice_tete is None:
        pass
    else:
        st.write("已上传发票抬头信息 : ", invoice_tete.name)
        datasender = pd.read_excel(invoice_tete)
        hide_table_row_index = """
                    <style>
                    thead tr th:first-child {display:none}
                    tbody th {display:none}
                    </style>
                    """
        st.markdown(hide_table_row_index, unsafe_allow_html=True)
        st.write(" Please find All sender information ")
        st.table(datasender[["发件人代码", "发件人英文"]])
        col1, col2 = st.columns([2, 9.5])
        with col1:
            option = st.number_input("Choose the sender:", min_value=0, max_value=len(datasender), key=int)
        with col2:
            if option <= 0 or option >= len(datasender) + 1:
                pass
            else:
                choose_sender = datasender["发件人英文"].loc[datasender["发件人代码"] == int(option)].tolist()[0]
                Nameofsender = choose_sender
                sender_adresse_complete = datasender["完整地址"].loc[datasender["发件人代码"] == int(option)].tolist()[0]
                Sendercountrycode = datasender["国家代码"].loc[datasender["发件人代码"] == int(option)].tolist()[0]
                Streetsender = datasender["地址"].loc[datasender["发件人代码"] == int(option)].tolist()[0]
                Citysender = datasender["城市"].loc[datasender["发件人代码"] == int(option)].tolist()[0]
                Senderzipcode = str(datasender["邮编"].loc[datasender["发件人代码"] == int(option)].tolist()[0]).split(".")[0]
                st.write("")
                st.write("You have choosen :", choose_sender)
                st.write("Adresse complet is : ", sender_adresse_complete)
        st.write("")
        st.subheader("第二步骤：上传清关发票数据，可批量上传")
        custom_invoice_datas = st.file_uploader("", type=(["xlsx", "xls"]),
                                                accept_multiple_files=True)
        if custom_invoice_datas is None:
            pass
        else:
            p = 0
            for custom_invoice_data in custom_invoice_datas:
                p = p + 1
                st.write("")
                st.write(" ##### :point_right:    处理第", str(p), "份清关材料 : ", custom_invoice_data.name)
                st.write(" - 请输入 提单总重, 包裹总数, 境内运费, 国际运费 (*为必填)")
                col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
                with col1:
                    lta_officel_weight_kg = st.number_input("*请输入提单毛重 : ", min_value=0, max_value=1000000, key=float)
                with col2:
                    lta_officel_carton = st.number_input("*请输入提单包裹数量 : ", min_value=0, max_value=1000000, key=int)
                with col3:
                    transport_fee_interne = st.number_input("*请输入欧洲境内运费 :", min_value=0, max_value=1000000, key=float)
                with col4:
                    transport_fee_externe = st.number_input("请输入国际运费: ", min_value=0, max_value=1000000, key=float)

                datainvoice = pd.read_excel(custom_invoice_data)
                datainvoice = datainvoice.dropna(subset=["货箱编号"])
                # 已有产品申报单价
                datainvoice['产品申报单价'] = datainvoice['产品申报单价'].apply(lambda x: float(x))
                datainvoice['产品申报数量'] = datainvoice['产品申报数量'].apply(lambda x: int(x))
                datainvoice['货箱重量(KG)'] = datainvoice['货箱重量(KG)'].apply(lambda x: float(x))
                datainvoice['跟踪号'] = datainvoice['跟踪号'].apply(lambda x: str(x).split(".")[0])
                datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: str(x)[:10])
                datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: int(x))
                datainvoice['申报总价'] = datainvoice['产品申报单价'] * datainvoice['产品申报数量']
                datainvoice['毛重比例'] = datainvoice['货箱重量(KG)'] / datainvoice['货箱重量(KG)'].sum()
                datainvoice['包裹净重'] = datainvoice['货箱重量(KG)'] - len(set(datainvoice['货箱编号'].tolist())) * 1 * \
                                      datainvoice['毛重比例']
                datainvoice['产品净重'] = ((datainvoice['包裹净重'] / datainvoice['产品申报数量']) - 0.005).round(2)
                datainvoice['包裹净重'] = round(datainvoice['产品净重'] * datainvoice['产品申报数量'], 2)
                datainvoice['箱数'] = datainvoice['货箱编号']  # 先等于运单号，然后在调整
                datainvoice['每公斤价值'] = round(datainvoice['申报总价'] / datainvoice['货箱重量(KG)'], 2)  # 先等于运单号，然后在调整
                value_total = datainvoice['申报总价'].sum()
                datainvoice['产品英文品名'] = datainvoice['产品英文品名']
                datainvoice['产品中文品名'] = datainvoice['产品中文品名']
                datainvoice = datainvoice.sort_values("货箱编号")
                datainvoice = datainvoice.fillna("")
                vats = list(set(datainvoice["VAT号"].tolist()))
                vats.sort()
                ltas = list(set(datainvoice["提单号"].tolist()))
                if len(ltas) == 1:
                    lta = ltas[0]
                else:
                    lta = str(ltas)
                kg_brut_total = datainvoice['货箱重量(KG)'].sum().round(2)
                carton_total = len(set(datainvoice['货箱编号'].tolist()))

                col1, col2 = st.columns([5, 5])
                with col1:
                    st.write(" ###### 请根据不同的业务，请选择对应的清关行：")
                    option = st.selectbox(
                        '',
                        ('','SMDG Logistics SRL',''))
                with col2:
                    template = st.file_uploader("上传对应清关模板")
                if st.button('生成清关材料👈'):
                    if value_total >= 50000:
                        st.write(':+1:申报价值高于5万欧',value_total, "   Euros    ")
                    else:
                        st.write(':triumph:申报总价值可能过低', "   ---申报价值：", str(value_total), "   Euros   ")

                    if lta_officel_weight_kg == kg_brut_total:
                        st.write(':+1:重量相符')
                    else:
                        st.write(':triumph:重量不符合', "   --- 输入重量：", str(lta_officel_weight_kg), "   KG    ,  ", "发票重量：",
                                 str(kg_brut_total), "   KG")
                    if lta_officel_carton == carton_total:
                        st.write(':+1:包裹数量相符')
                    else:
                        st.write(':triumph:包裹数量不符合', "   --- 输入包裹数量：", str(lta_officel_carton), "   Cartons    ,  ",
                                 "发票包裹数量：", str(carton_total), "   Cartons")
                    st.write("")
                    if transport_fee_interne == 0:
                        transport_fee_interne = 1300
                    if transport_fee_externe == 0:
                        transport_fee_externe = ""
                    col1, col2, col3 = st.columns([1, 5, 5])
                    with col1:
                        pass
                    with col2:
                        st.write("- ###### 提单号码    : ", lta)
                        st.write("- ###### 清关材料数量 : ", str(len(vats)), " Docs")
                    with col3:
                        st.write("- ###### 包裹总数       : ", str(carton_total), " Cartons")
                        st.write("- ###### 包裹总重       : ", str(kg_brut_total), " KG")
                    # 开始生成清关材料：
                    st.write("")
                    st.write('##### 您已选择', option, " 作为清关服务商")
                    st.write(':punch: SMDG将为您提供合理的清关价格,SMDG将为您提供和更高效的服务')
                    if template is None:
                        st.write(':punch: 请重新选择清关行或者上传清关模板')
                    else:
                        st.write(template.name)
                        if option == "SMDG Logistics SRLL":
                            st.write(" - 感谢您的信任，SMDG 正在筹备清关资质，预计2023年年初可以开始独立自主的清关业务")
                            st.write(" - 进一步消息请联系 邮箱 ： info@smdg.eu")
                            st.write(" - :pray:请重新选择清关行. 为带来不便, 深感抱歉")
                        elif option == "Cacesa":
                            st.write(" - 清关材料完善中...")
                            st.write(" - :pray:为带来不便, 深感抱歉")
                        elif option == "Flying":
                            st.write(" - 清关材料完善中...")
                            st.write(" - :pray:为带来不便, 深感抱歉")
                        elif option == "ECLL":
                            st.write(" - 清关材料完善中...")
                            st.write(" - :pray:为带来不便, 深感抱歉")
                        elif option == "SMDG Logistics SRL":
                            zip_file_name = str(lta) + 'CI+PL+Manifest.zip'
                            zip_file = zipfile.ZipFile(zip_file_name, 'w')
                            dic_lta = []
                            a = 0
                            align = Alignment(horizontal='left', vertical='center')
                            side = Side(style='thin', color='000000')
                            border = Border(top=side, bottom=side, left=side, right=side)
                            for vat in vats:
                                a = a + 1
                                datainvoice_vat = datainvoice.loc[datainvoice['VAT号'] == vat]
                                # 获取交货条款;交货城市;清关方式;收件人国家
                                incoterme = list(set(datainvoice_vat["交货条款"].tolist()))[0]
                                incoterme_city = list(set(datainvoice_vat["交货城市"].tolist()))[0]
                                delivery_country = list(set(datainvoice_vat["收件人国家"].tolist()))[0]
                                code_regime = list(set(datainvoice_vat["清关方式"].tolist()))[0]
                                qty_carton = len(set(datainvoice_vat["货箱编号"].tolist()))
                                exporter_chi = "---"
                                exporter_eng = Nameofsender
                                ref_number = lta + " - " + str(a)
                                invoice_number = "HBL - " + lta + " - " + str(a)
                                importer = datainvoice_vat["收件人"].tolist()[0]
                                EORI = datainvoice_vat["EORI"].tolist()[0]
                                adresse = datainvoice_vat["地址"].tolist()[0]
                                code_postal = str(datainvoice_vat["邮编"].tolist()[0]).split(".")[0]
                                city = datainvoice_vat["城市"].tolist()[0]
                                county_2_chiffre = datainvoice_vat["国家代码"].tolist()[0]
                                county_complet = datainvoice_vat["国家全称"].tolist()[0]
                                adresse_importer_complet = str(adresse) + " ," + str(
                                    code_postal) + " ," + str(city) + " ," + str(
                                    county_complet)

                                # 填写文件
                                target = lta + " - INV&PL- " + vat + " - " + str(qty_carton) + "pcs ( HBL " + str(
                                    a) + ").xlsx"
                                st.write(target)
                                wb = load_workbook(template)
                                invoice_sheet = wb.worksheets[0]
                                # 填写excel invoice 表头信息
                                invoice_sheet.cell(1, 1, exporter_chi)  # 出口商公司
                                invoice_sheet.cell(2, 1, exporter_eng)  # 出口商英文
                                invoice_sheet.cell(4, 3, exporter_eng)  # 出口商英文
                                invoice_sheet.cell(5, 3, ref_number)  # 分单号
                                invoice_sheet.cell(6, 3, Streetsender)  # 地址
                                invoice_sheet.cell(7, 3, str(Senderzipcode))  # 邮编
                                invoice_sheet.cell(8, 3, Citysender)  # 邮编
                                invoice_sheet.cell(9, 3, Sendercountrycode)  # 邮编
                                invoice_sheet.cell(4, 10, invoice_number)  # 发票号码
                                invoice_sheet.cell(5, 10, date_now)  # 发日期

                                # 填写进口商信息 excel invoice 表头信息
                                invoice_sheet.cell(11, 3, importer)  # 进口商公司名称
                                invoice_sheet.cell(12, 3, "")  # 电话
                                invoice_sheet.cell(13, 3, adresse)  # 地址
                                invoice_sheet.cell(14, 3, code_postal)  # 邮编
                                invoice_sheet.cell(15, 3, city)  # 城市
                                invoice_sheet.cell(16, 3, county_2_chiffre)  # 国家
                                invoice_sheet.cell(17, 3, delivery_country)  # 收货国家
                                invoice_sheet.cell(11, 10, vat)
                                invoice_sheet.cell(12, 10, EORI)
                                invoice_sheet.cell(13, 10, "EUR")  # 币种
                                invoice_sheet.cell(14, 10, incoterme)
                                invoice_sheet.cell(15, 10, incoterme_city)
                                invoice_sheet.cell(16, 10, code_regime)  # 递延
                                invoice_sheet.cell(17, 10, "")  # 邮箱
                                wb.save(target)
                                # 填写excel invoice 主体信息
                                datainvoice_vat_traiter = datainvoice_vat[
                                    ["产品英文品名", "产品海关编码", '产品申报单价', '产品中文品名', '材质（须填写英文）', '货箱编号', '产品申报数量', '申报总价',
                                     '包裹净重', '货箱重量(KG)', "产品销售链接",
                                     "运单号"]]
                                for x in range(len(datainvoice_vat_traiter)):
                                    for y in range(0, 12):
                                        column = datainvoice_vat_traiter.columns[y]
                                        line = 20 + int(x)
                                        valeur = datainvoice_vat_traiter[column].tolist()[x]
                                        invoice_sheet.cell(line, y + 2, valeur)
                                        invoice_sheet.cell(line, y + 2).border = border
                                        invoice_sheet.cell(line, y + 2).alignment = align
                                    # 合并单元格
                                # marks_list = []
                                # tracking_list = []
                                shipement_list = []
                                for row in range(20, line + 1):
                                    # marks = invoice_sheet['G' + str(row)].value
                                    shipement = invoice_sheet['M' + str(row)].value
                                    # marks_list.append(marks)
                                    shipement_list.append(shipement)
                                # 调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作
                                start_row = 20  # 开始行是第20行
                                # Merge_cells(invoice_sheet, marks_list, start_row, "G")
                                Merge_cells(invoice_sheet, shipement_list, start_row, "M")  # "M" - 票在最后一列
                                # 填写excel invoice 结尾西信息
                                invoice_sheet.delete_rows(line + 1, 2000 - line - 1)
                                invoice_sheet.merge_cells(start_row=line + 1, start_column=3, end_row=line + 1,
                                                          end_column=5)
                                sum_pcs = datainvoice_vat_traiter['产品申报数量'].sum()
                                invoice_sheet.cell(line + 1, 8, sum_pcs)
                                sum_total_value = datainvoice_vat_traiter['申报总价'].sum()
                                invoice_sheet.cell(line + 1, 9, sum_total_value)
                                sum_total_net = datainvoice_vat_traiter['包裹净重'].sum()
                                invoice_sheet.cell(line + 1, 10, sum_total_net)
                                sum_total_brut = datainvoice_vat_traiter['货箱重量(KG)'].sum()
                                invoice_sheet.cell(line + 1, 11, sum_total_brut)
                                invoice_sheet.cell(line + 6, 3,
                                                   round(transport_fee_interne * (sum_total_brut / kg_brut_total)))

                                wb.save(target)
                                # 处理 packing list
                                pl_sheet = wb.worksheets[1]
                                pl_sheet.cell(1, 1, exporter_chi)  # 发票抬头
                                pl_sheet.cell(2, 1, exporter_eng)  # 发票英文名称
                                pl_sheet.cell(4, 2, invoice_number)  # 发票英文名称
                                data_pl_vat_traiter = datainvoice_vat[
                                    ["产品英文品名", '产品申报单价', '产品中文品名', '货箱编号', '产品净重', '箱数', '产品申报数量', '包裹净重', '货箱重量(KG)']]
                                for x in range(len(data_pl_vat_traiter)):
                                    for y in range(9):
                                        column = data_pl_vat_traiter.columns[y]
                                        line = 6 + int(x)
                                        valeur = data_pl_vat_traiter[column].tolist()[x]
                                        pl_sheet.cell(line, y + 2, valeur)
                                        pl_sheet.cell(line, y + 2).border = border
                                        pl_sheet.cell(line, y + 2).alignment = align
                                # 合单元格
                                # marks_list = []  # 唛头
                                carton_list = []  # 包裹数量
                                for row in range(6, line + 1):
                                    # marks = pl_sheet['E' + str(row)].value
                                    carton = pl_sheet['G' + str(row)].value
                                    # marks_list.append(marks)
                                    carton_list.append(carton)
                                # 调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作
                                start_row = 6  # 开始行是第六行
                                # Merge_cells(pl_sheet, marks_list, start_row, "E")  # "E" - 唛头是a列
                                Merge_cells(pl_sheet, carton_list, start_row, "G")  # "G" - 箱数是在C列
                                pl_sheet.delete_rows(line + 1, 2000 - line - 1)
                                pl_sheet.cell(line + 1, 8, sum_pcs)
                                pl_sheet.cell(line + 1, 7, qty_carton)
                                pl_sheet.cell(line + 1, 9, sum_total_net)
                                pl_sheet.cell(line + 1, 10, sum_total_brut)
                                for row in range(6, line + 1):
                                    try:
                                        value = str(pl_sheet.cell(row, 7).value)
                                        if len(value) > 6:
                                            pl_sheet.cell(row, 7, 1)
                                        else:
                                            pl_sheet.cell(row, 7, 0)
                                    except:
                                        pass
                                wb.save(target)

                                # 处理 resume
                                resume_sheet = wb.worksheets[2]
                                datainvoice_vat_resume = datainvoice_vat_traiter.groupby(by="产品海关编码", sort=True).sum()
                                nb_hscode = len(datainvoice_vat_resume)
                                descriptions_hbl = list(set(datainvoice_vat_traiter["产品英文品名"].tolist()))
                                descriptions_hbl.sort()
                                descriptions_hbl = str(descriptions_hbl).replace("{", '').replace("}", '').replace("'",
                                                                                                                   '').replace(
                                    "[", '').replace("]", '')
                                for x in range(len(datainvoice_vat_resume)):
                                    data = datainvoice_vat_resume[x:x + 1]
                                    hscode = data.index[0]
                                    descriptions = list(set(datainvoice_vat_traiter["产品英文品名"].loc[
                                                                datainvoice_vat_traiter["产品海关编码"] == hscode].tolist()))
                                    descriptions.sort()
                                    descriptions = str(descriptions).replace("{", '').replace("}", '').replace("'",
                                                                                                               '').replace(
                                        "[", '').replace("]", '')
                                    qty_hscode = data["产品申报数量"].tolist()[0]
                                    value_hscode = data["申报总价"].tolist()[0]
                                    kgnet_hscode = data["包裹净重"].tolist()[0]
                                    kgbrut_hscode = data["货箱重量(KG)"].tolist()[0]
                                    resume_sheet.cell(x + 2, 1, str(hscode))
                                    resume_sheet.cell(x + 2, 1).border = border
                                    resume_sheet.cell(x + 2, 2, qty_hscode)
                                    resume_sheet.cell(x + 2, 2).border = border
                                    resume_sheet.cell(x + 2, 3, value_hscode)
                                    resume_sheet.cell(x + 2, 3).border = border
                                    resume_sheet.cell(x + 2, 4, kgnet_hscode)
                                    resume_sheet.cell(x + 2, 4).border = border
                                    resume_sheet.cell(x + 2, 5, kgbrut_hscode)
                                    resume_sheet.cell(x + 2, 5).border = border
                                    resume_sheet.cell(x + 2, 6, descriptions)
                                    resume_sheet.cell(x + 2, 6).border = border

                                resume_sheet.delete_rows(len(datainvoice_vat_resume) + 2,
                                                         1000 - len(datainvoice_vat_resume) - 2)
                                resume_sheet.cell(x + 3, 2, datainvoice_vat_traiter["产品申报数量"].sum())
                                resume_sheet.cell(x + 3, 3, datainvoice_vat_traiter["申报总价"].sum())
                                resume_sheet.cell(x + 3, 4, datainvoice_vat_traiter["包裹净重"].sum())
                                resume_sheet.cell(x + 3, 5, datainvoice_vat_traiter["货箱重量(KG)"].sum())
                                wb.save(target)

                                def get_binary_file_downloader_html(bin_file, file_label='File'):
                                    with open(file_path, 'rb') as f:
                                        data = f.read()
                                    bin_str = base64.b64encode(data).decode()
                                    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">点击下载 {file_label}</a>'
                                    return href

                                file_path = target
                                file_label = target
                                st.markdown(get_binary_file_downloader_html(file_path, file_label),
                                            unsafe_allow_html=True)

                                dic_resume = {"提单": lta,
                                              "分单": lta + "-HBL-" + str(a),
                                              "税号": vat,
                                              "包裹数量": qty_carton,
                                              "净重": sum_total_net,
                                              "毛重": sum_total_brut,
                                              "海关码数量": nb_hscode,
                                              "申报金额": sum_total_value,
                                              "Description": descriptions_hbl,
                                              "Company Trading": Nameofsender,
                                              "Adresse shiper 1": Streetsender,
                                              "Adresse shiper 2": str(Citysender) + " " + str(
                                                  Senderzipcode) + " " + str(Sendercountrycode),
                                              "Consignee": importer,
                                              "Adresse Consignee 1": adresse,
                                              "Adresse Consignee 2": str(city) + " " + str(code_postal) + " " + str(county_complet),
                                              "CBM": "",
                                              "Place of recepit": "",
                                              "Port of loading": "",
                                              "Ocean Vessel": "",
                                              "Port of discharge": "",
                                              "SealNo": "",
                                              "Type": "",
                                              "Rate": "",
                                              "Prepaid at": "",
                                              "Place Date Issu": ""}
                                dic_lta.append(dic_resume)

                                # 开始生成 Begate文件
                                dic_file = []
                                list_hscode = list(set(datainvoice_vat["产品海关编码"].tolist()))
                                list_hscode.sort()
                                op = 1
                                for hscode in list_hscode:
                                    description = set(
                                        datainvoice_vat["产品英文品名"].loc[datainvoice_vat["产品海关编码"] == hscode].tolist())
                                    Gooddescription = str(description).replace("{", '').replace("}", '').replace("'",
                                                                                                                 '')
                                    Typeofpackages = "PC"
                                    Numberofpackages = datainvoice_vat["产品申报数量"].loc[
                                        datainvoice_vat["产品海关编码"] == hscode].sum()
                                    Brand_Marks = ""
                                    Netweight = datainvoice_vat["包裹净重"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
                                    Grossweight = datainvoice_vat["货箱重量(KG)"].loc[
                                        datainvoice_vat["产品海关编码"] == hscode].sum()
                                    Value = datainvoice_vat["申报总价"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
                                    Countryoforigin = "CN"
                                    Nameofsender = Nameofsender
                                    Streetsender = Streetsender
                                    Citysender = Citysender
                                    Senderzipcode = Senderzipcode
                                    Sendercountrycode = Sendercountrycode
                                    EORIsender = ""
                                    Nameofconsignee = importer
                                    Streetconsignee = adresse
                                    Cityofconsignee = city
                                    Zipcodeconsignee = code_postal
                                    Countrycodeconsignee = county_2_chiffre
                                    Track_Trace = op
                                    op = op + 1
                                    codeadditionnel = ""
                                    Invoicecurrency = "EUR"
                                    Incoterm = incoterme  # 这里注意
                                    countrycodeofdestination = delivery_country
                                    consigneeID = EORI
                                    if "GVR" in str(code_regime):
                                        code_regime = "4000"
                                    else:
                                        code_regime = code_regime
                                    or_4000_4200 = code_regime
                                    dic_hscode = {"HSCode": hscode,
                                                  "Gooddescription": Gooddescription,
                                                  "Typeofpackages": Typeofpackages,
                                                  "Numberofpackages": Numberofpackages,
                                                  "Brand_Marks": Brand_Marks,
                                                  "Netweight": Netweight,
                                                  "Grossweight": Grossweight,
                                                  "Value": Value,
                                                  "Countryoforigin": Countryoforigin,
                                                  "Nameofsender": Nameofsender,
                                                  "Streetsender": Streetsender,
                                                  "Citysender": Citysender,
                                                  "Senderzipcode": Senderzipcode,
                                                  "Sendercountrycode": Sendercountrycode,
                                                  "EORIsender": EORIsender,
                                                  "Nameofconsignee": Nameofconsignee,
                                                  "Streetconsignee": Streetconsignee,
                                                  "Cityofconsignee": Cityofconsignee,
                                                  "Zipcodeconsignee": Zipcodeconsignee,
                                                  "Countrycodeconsignee": Countrycodeconsignee,
                                                  "Track_Trace": Track_Trace,
                                                  "codeadditionnel": codeadditionnel,
                                                  "Invoicecurrency": Invoicecurrency,
                                                  "Incoterm": Incoterm,
                                                  "countrycodeofdestination": delivery_country,
                                                  "consigneeID": consigneeID,
                                                  "or_4000_4200": or_4000_4200}
                                    dic_file.append(dic_hscode)
                                df_begate = pd.DataFrame(list(dic_file))
                                begate_name = lta + " - BEGATE- " + vat + " (HBL " + str(a) + ").xlsx"
                                df_begate.to_excel(begate_name, sheet_name='Begate file', index=False)

                                def get_binary_file_downloader_html(bin_file, file_label='File'):
                                    with open(file_path, 'rb') as f:
                                        data = f.read()
                                    bin_str = base64.b64encode(data).decode()
                                    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">点击下载 {file_label}</a>'
                                    return href

                                file_path = begate_name
                                file_label = begate_name
                                st.markdown(get_binary_file_downloader_html(file_path, file_label),
                                            unsafe_allow_html=True)

                            dic_resume = {"提单": "共计",
                                          "分单": "",
                                          "税号": "",
                                          "包裹数量": len(set(datainvoice["货箱编号"].tolist())),
                                          "净重": datainvoice["包裹净重"].sum(),
                                          "毛重": datainvoice["货箱重量(KG)"].sum(),
                                          "海关码数量": "",
                                          "申报金额": datainvoice["申报总价"].sum()}
                            dic_lta.append(dic_resume)
                            df_lta = pd.DataFrame(list(dic_lta))
                            df_lta_name = lta + " 税号信息总结.xlsx"
                            df_lta.to_excel(df_lta_name, sheet_name='税号信息总结', index=False)
                            file_path = df_lta_name
                            file_label = df_lta_name
                            st.markdown(get_binary_file_downloader_html(file_path, file_label),
                                        unsafe_allow_html=True)


def declaration_product(product):
    payload = {"includeUK": "false",
               "lang": "CN",
               "partner": "CN",
               "product": product,
               "years": '2021'}
    headers = {'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/102.0.0.0 Safari/537.36',
               'Cookie': ''}
    r = requests.get(
        'https://webgate.ec.europa.eu/flows/public/v1/stats?', params=payload, headers=headers)
    list_value = r.json()['rows']
    importValue_total, importQuantity_total = 0, 0
    for value in list_value:
        country = value['country']
        importValue = value['samples']['2021']['importValue']
        if len(str(importValue).split(".")[-1]) == 3:
            importValue = int(str(importValue).replace(".", ""))
        elif len(str(importValue).split(".")[-1]) == 2:
            importValue = int(str(importValue).replace(".", "")) * 10
        elif len(str(importValue).split(".")[-1]) == 1:
            importValue = int(str(importValue).replace(".", "")) * 100

        importQuantity = value['samples']['2021']['importQuantity']
        if len(str(importQuantity).split(".")[-1]) == 3:
            importQuantity = int(str(importQuantity).replace(".", ""))
        elif len(str(importQuantity).split(".")[-1]) == 2:
            importQuantity = int(str(importQuantity).replace(".", "")) * 10
        elif len(str(importQuantity).split(".")[-1]) == 1:
            importQuantity = int(str(importQuantity).replace(".", "")) * 100

        importValue_total, importQuantity_total = \
            importValue_total + importValue, importQuantity_total + importQuantity
    country = "EURO 27"
    if importQuantity_total == 0:
        import_kg_total = 0
    else:
        import_kg_total = round(importValue_total / importQuantity_total, 2)
    return import_kg_total


def Merge_cells(ws, target_list, start_row, col):  # 合并单元格
    '''
    ws: 是需要操作的工作表
    start_row: 是开始行，即工作表中开始比对数据的行（需要将标题除开）
    col: 是需要处理数据的列
    '''
    start = 0  # 开始行计数，初试值为0，对应列表中的第1个元素的位置0
    end = 0  # 结束行计数，初试值为0，对应列表中的第1个元素的位置0
    reference = target_list[0]  # 设定基准，以列表中的第一个字符串开始
    for i in range(len(target_list)):  # 遍历列表
        if target_list[i] != reference:  # 开始比对，如果内容不同执行如下
            reference = target_list[i]  # 基准变成列表中下一个字符串
            end = i - 1  # 列计数器
            ws.merge_cells(col + str(start + start_row) + ":" + col + str(end + start_row))
            start = end + 1
        if i == len(target_list) - 1:  # 遍历到最后一行，按如下操作
            end = i
            ws.merge_cells(col + str(start + start_row) + ":" + col + str(end + start_row))


def decision(a):
    if (len(str(a)) == 0):
        return ''
    elif (a >= 0):
        return '无'
    elif a < 0:
        return '有'


def get_invoicedate(source):
    writer_1 = pd.ExcelFile(source)
    c = writer_1.sheet_names
    datainvoice = writer_1.parse(c[0])
    datainvoice = datainvoice.dropna(subset=["货箱编号"])
    # 已有产品申报单价
    datainvoice['产品申报单价'] = datainvoice['产品申报单价'].apply(lambda x: float(x))
    datainvoice['产品申报数量'] = datainvoice['产品申报数量'].apply(lambda x: int(x))
    datainvoice['货箱重量(KG)'] = datainvoice['货箱重量(KG)'].apply(lambda x: float(x))
    datainvoice['跟踪号'] = datainvoice['跟踪号'].apply(lambda x: str(x).split(".")[0])
    datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: str(x)[:10])
    datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: int(x))
    datainvoice['申报总价'] = datainvoice['产品申报单价'] * datainvoice['产品申报数量']

    datainvoice['毛重比例'] = datainvoice['货箱重量(KG)'] / datainvoice['货箱重量(KG)'].sum()
    datainvoice['包裹净重'] = datainvoice['货箱重量(KG)'] - len(set(datainvoice['货箱编号'].tolist())) * 1 * datainvoice['毛重比例']
    datainvoice['产品净重'] = ((datainvoice['包裹净重'] / datainvoice['产品申报数量']) - 0.005).round(2)
    datainvoice['包裹净重'] = round(datainvoice['产品净重'] * datainvoice['产品申报数量'], 2)

    datainvoice['箱数'] = datainvoice['货箱编号']  # 先等于运单号，然后在调整
    datainvoice['每公斤价值'] = round(datainvoice['申报总价'] / datainvoice['货箱重量(KG)'], 2)  # 先等于运单号，然后在调整
    datainvoice['产品英文品名'] = datainvoice['产品英文品名']
    datainvoice['产品中文品名'] = datainvoice['产品中文品名']
    datainvoice = datainvoice.sort_values("货箱编号")
    datainvoice = datainvoice.fillna("")
    return datainvoice


def study_invoice(data_hscode, source):
    today = date.today()
    years = 2021
    datainvoice = get_invoicedate(source)
    df_hscode_invoice = datainvoice[
        ["运单号", "申报总价", "产品海关编码", "每公斤价值", "产品英文品名", "产品中文品名"]].drop_duplicates().sort_values("产品海关编码")
    df_hscode_analyse = pd.merge(df_hscode_invoice, data_hscode, left_on="产品海关编码", right_on="hscode", how='left')
    list_hscode_no_info = set(df_hscode_analyse['产品海关编码'].loc[df_hscode_analyse['hscode'].isna()].tolist())
    if len(list_hscode_no_info) == 0:
        pass
    else:
        list_o = []
        hscode_no_exsite = []
        n = 0
        st.write("共计%s个海关码不再数据库，需进行海关网站抓取" % (len(list_hscode_no_info)))
        print("-------------------------")
        for hscode_on_info in list_hscode_no_info:
            hscode_on_info = str(hscode_on_info)[:10]
            n = n + 1
            st.write("正在提取%s个海关码 :" % (n), hscode_on_info)
            if len(str(hscode_on_info)) == 10:
                try:
                    description_hscode, anti_dumping, duty = extrait_hscode(hscode_on_info, today)
                    description_en_chinois = "" #translate_eng_cn(description_hscode)
                    product = str(hscode_on_info)[:8]
                    import_kg_total = declaration_product(product)
                    a = {'hscode': hscode_on_info, 'Duty': duty, 'import_euro_kg': import_kg_total,
                         'anti_dumping': anti_dumping, 'description_hscode': description_hscode,
                         'description_en_chinois': description_en_chinois,
                         'date_search': today, 'lien': ''}
                    list_o.append(a)
                    st.write("****************************海关码存在，已缓存  %s   ，" % (hscode_on_info))

                except:
                    b = {'hscode': hscode_on_info, 'Statue': "未找到，人工核实"}
                    hscode_no_exsite.append(b)
                    st.write("****************************未找到海关码  %s   ，请核实" % (hscode_on_info))
            else:
                st.write("海关码为10位数，请补充完整")

        df_no_existe = pd.DataFrame(list(hscode_no_exsite))
        df_hscode_insert = pd.DataFrame(list(list_o))
        data_hscode = data_hscode.append(df_hscode_insert, ignore_index=True)
    df_hscode_analyse = pd.merge(df_hscode_invoice, data_hscode, left_on="产品海关编码", right_on="hscode", how='left')
    df_hscode_analyse["差值"] = df_hscode_analyse["每公斤价值"] - df_hscode_analyse["import_euro_kg"]
    df_hscode_analyse["低报风险"] = df_hscode_analyse['差值'].apply(decision)
    df_antidumping = df_hscode_analyse[df_hscode_analyse["anti-dumping"] == "anti-dumping"]
    df_low_value = df_hscode_analyse[df_hscode_analyse["低报风险"] == "有"]
    table_df_low_value = pd.pivot_table(df_low_value, values=['import_euro_kg', '每公斤价值', '差值'],
                                        index=['产品海关编码', '低报风险', '产品英文品名', '产品中文品名'],
                                        aggfunc={'import_euro_kg': np.mean,
                                                 '每公斤价值': np.mean,
                                                 '差值': np.mean})
    table = pd.pivot_table(df_hscode_analyse, values=['import_euro_kg', '每公斤价值'],
                           index=['产品海关编码', 'description_en_chinois', '产品中文品名'],
                           aggfunc={'import_euro_kg': np.mean,
                                    '每公斤价值': np.mean})
    with pd.ExcelWriter("清关文件海关码分析结果.xlsx", engine="openpyxl") as writer:
        df_hscode_analyse.to_excel(writer, sheet_name='申报信息总结', index=False)
        table.to_excel(writer, sheet_name='透视表格')
        try:
            df_no_existe.to_excel(writer, sheet_name='海关码不存在', index=False)
        except:
            pass
        df_antidumping.to_excel(writer, sheet_name='反倾销', index=False)
        try:
            table_df_low_value.to_excel(writer, sheet_name='低报风险')
        except:
            pass
        try:
            df_hscode_insert.to_excel(writer, sheet_name='打包发给米西', index=False)
        except:
            pass

    def get_binary_file_downloader_html(bin_file, file_label='File'):
        with open(file_path, 'rb') as f:
            data = f.read()
        bin_str = base64.b64encode(data).decode()
        href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">点击下载 {file_label}</a>'
        return href

    file_path = "清关文件海关码分析结果.xlsx"
    file_label = "清关文件海关码分析结果"
    st.markdown(get_binary_file_downloader_html(file_path, file_label),
                unsafe_allow_html=True)


def hs_code():
    import streamlit as st
    import pandas as pd
    import pydeck as pdk
    from urllib.error import URLError
    st.markdown(f"# SMDG{list(page_names_to_funcs.keys())[2]}" + "智能化服务")

    st.write(
        """ 在这里您可以获得不一样的服务 """
    )
    option = st.selectbox("请选择海关码服务", ["", "海关码查询服务", "清关数据检查", "税金预估"])
    if option == "海关码查询服务":
        hscodes = st.text_input("请输入海关码：备注海海关码之间已 ' , ' 隔开")
        if len(hscodes) != 0:
            hscodes = hscodes.replace("，", ",")
            hscodes = list(set(str(hscodes).split(",")))
            st.text("查询：" + str(hscodes))
            for hscode in hscodes:
                a = 5
                if a == 5:
                    dic = []
                    url = "https://eservices.minfin.fgov.be/extTariffBrowser/Measure?cnCode=%s&country=29422&trade=0&cssfile=tarbro" \
                          "&date=%s&lang=EN&page=1" % (
                              hscode, today)
                    headers = {
                        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) '
                                      'Chrome/76.0.3809.132 Safari/537.36'}
                    res = requests.get(url, headers=headers)
                    soup = BeautifulSoup(res.text, 'html.parser')
                    hscode = soup.find('span', class_="smaller-title").text.replace(" ", "")
                    description_hscode = soup.find('ul', class_="nostyle").getText().replace('\n', '').replace(
                        '                                   ', ' /')
                    footnote = soup.find('table', class_="table-nopadding").getText().replace('\n', '').strip().replace(
                        "Footnotes:",
                        "")
                    supplementary_unit = soup.find('table', class_="table-nopadding bottom-aligned").getText().replace(
                        '\n',
                        '').strip().replace(
                        "Supplementary unit:", "")
                    tables = soup.find_all('div', class_="meas-header")
                    for table in tables:
                        table_infos = table.getText().split("\n")
                        type_table = table_infos[0]  # 表格类型
                        table_infos_sorts = table_infos[19:]
                        nb_ligne = len(table_infos_sorts) / 20
                        for x in range(int(nb_ligne)):
                            Geographical_area = table_infos_sorts[20 * x + 0] + "  " + table_infos_sorts[20 * x + 1]
                            Measure_type = table_infos_sorts[20 * x + 2] + "  " + table_infos_sorts[20 * x + 3]
                            Tariff = table_infos_sorts[20 * x + 4] + "  " + table_infos_sorts[20 * x + 5]
                            dic_0 = {"type_table": type_table,
                                     "Measure_type": Measure_type,
                                     "Tariff": Tariff,
                                     "Geographical_area": Geographical_area, }
                            dic.append(dic_0)
                    pd_hscode_no_info = pd.DataFrame(list(dic))
                    if "CN - China  " in pd_hscode_no_info["Geographical_area"].tolist():
                        anti_dumping = "anti-dumping"
                    else:
                        anti_dumping = "-"
                    duty = pd_hscode_no_info["Tariff"].loc[(pd_hscode_no_info["type_table"] == "Tariff measures") & (
                            pd_hscode_no_info["Measure_type"] == "Third country duty          ")].tolist()[0]
                    description_en_chinois = ""# translate_eng_cn(description_hscode)
                    try:
                        import_kg_total = declaration_product(hscode[:8])
                    except:
                        import_kg_total = 1000000000
                    df_hscode = pd.DataFrame([["海关码", hscode],
                                              ["海关关税", duty],
                                              ["反倾销", anti_dumping],
                                              ["2021申报", str(import_kg_total) + " €/KG"],
                                              ["英文解释", description_hscode],
                                              ["中文品名", description_en_chinois],
                                              ["补充单元", supplementary_unit],
                                              ["脚注", footnote]],
                                             columns=['DESCRIPTION', 'INFORMATION'])
                    st.table(df_hscode)
    elif option == "清关数据检查":
        source = st.file_uploader("上传清关资料", type=(["xlsx", "xls"]))
        path = "https://raw.githubusercontent.com/SMDGLogisticsSRl/web-service/70f75f3da2b92a37b292bed7ff2f9ed967ea10ec/hscode_database.txt"
        data_hscode = pd.read_table(path, sep='\t')
        if source is not None:
            line_resultat = study_invoice(data_hscode, source)


3


def air_pick_up():
    import streamlit as st
    import pandas as pd
    import altair as alt

    from urllib.error import URLError

    st.markdown(f"# SMDG {list(page_names_to_funcs.keys())[3]}业务")

    st.write(
        """
        **此模板为方便周末提货使用**
        \n 1. 发邮件给指定海关通知提货
             \n 2. 卡车公司订单
             \n 3. 货站信息""")
    options = st.selectbox("请选择服务", ["", "Transfert", "Truck Order", "Loading Instruction"])
    if options == "Transfert":
        st.write("准备邮件给海关")
        col1, col2, col3, col4 = st.columns([5, 5, 5, 5])
        with col1:
            handler = st.selectbox("选择货站", ('', 'AVIA', 'SWP', 'WFS', 'LACHS', 'BAS'))
            lta = st.text_input("输入提单号：")
            xuhao = st.text_input("输入邮件序号：")

        with col2:
            pcs = st.text_input("输入包裹数量：")
            kg = st.text_input("输入包裹重量：")

        with col3:
            date_noa_1 = st.date_input("输入NOA日期：")
            date_noa_2 = st.time_input("输入NOA时间：")

        with col4:
            date_pick_up_1 = st.date_input("输入提e货日期：")
            date_pick_up_2 = st.time_input("输入提货时间：")

        if st.button("准备并发送邮件"):
            date_noa = str(date_noa_1) + " " + str(date_noa_2)
            date_pick_up = str(date_pick_up_1) + " " + str(date_pick_up_2)

            dfges = get_data(handler, date_noa, date_pick_up, lta, pcs, kg)
            st.write("查看邮件内容模板")
            to_addrs = "fuqing.yuan@smdg.eu"
            title = "< N°%s > notification d’entrée en installation de stockage temporaire (TSD)(%s - SMDG)_LTA: %s " \
                    "TRANSFERT" % (xuhao, handler, lta)
            st.write("收件人：", to_addrs)
            st.write("邮件标题：", title)
            st.write(dfges)
            html = f"""
                           <!DOCTYPE html>
                           <head>
                           <style>
                                tr:nth-child(even) {{
                                background-color: #f2f2f2;
                                }}
                           </style>
                           </head>
                           <td>
                                Bonjour,Madame, Monsieur:
                           </td>
                           <ul>
                                L’envoi fera objet d’un transfert manifest .
                           </ul>
                           <body>
                                {dfges.to_html(index=False, escape=False)}
                           <p>
                                l'Equipe de SMDG Logistics SRL  .
                           </p>
                           <li>
                                Mes salutations distinguées 
                           </li>   
                           </body>
                           </html>"""

            html_msg = html
            msg = email.mime.multipart.MIMEMultipart()
            sender_show = 'fuqing.yuan@smdg.eu'
            recipient_show = 'fuqing.yuan.univ@gmail.com'
            cc_show = 'fuqing.yuan.univ@gmail.com'
            to_addrs = 'fuqing.yuan@smdg.eu'
            msg["Subject"] = title
            # 发件人显示，不起实际作用
            msg["from"] = sender_show
            # 收件人显示，不起实际作用
            msg["to"] = to_addrs
            # 抄送人显示，不起实际作用
            msg["Cc"] = cc_show
            msg.attach(MIMEText(html_msg, "html", "utf-8"))
            user = 'fuqing.yuan@smdg.eu'
            password = 'Beijing2008'
            with SMTP_SSL(host="smtp.exmail.qq.com", port=465) as smtp:
                smtp.login(user=user, password=password)
                smtp.sendmail(from_addr=user, to_addrs=to_addrs, msg=msg.as_string())
                st.success("邮件发送成功！")

    elif options == "Truck Order":
        st.write("发邮件卡车订单")
    elif options == "Loading Instruction":
        st.write("发邮件卡车订单")


page_names_to_funcs = {
    "公司介绍": intro,
    "清关资料": custom_invoice,
    "海关码": hs_code,
    "空运提货": air_pick_up
}

demo_name = st.sidebar.selectbox("请选择服务", page_names_to_funcs.keys())
page_names_to_funcs[demo_name]()
